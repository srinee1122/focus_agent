"""
sales_module.py — Sales Report agent logic (Phase 1: uploaded day books).

Responsibilities:
  - parse_sales_excel(): read a Focus "Sales day book" export into row dicts
  - upsert_sales(): store rows with (voucher, item) replace-dedupe
  - build_report(): salesman x item quantity matrix with targets/FOC
  - report_pdf(): render a report dict to a PDF file

Kept separate from main.py so the logic is testable standalone.
"""
from __future__ import annotations
import io
import re
import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd

# ────────────────────────────────────────────────────────────────────
# Parsing
# ────────────────────────────────────────────────────────────────────

# Header names as they appear in the Focus export (row with "Date")
_COLMAP = {
    "date":          ["date"],
    "voucher":       ["voucher"],
    "salesman":      ["salesman name", "salesman"],
    "customer":      ["customer account name", "customer"],
    "item_name":     ["item"],
    "qty":           ["quantity"],
    "qty_pieces":    ["quantity in base unit"],
    "unit":          ["unit"],
    "rate":          ["rate"],
    "rate_pcs":      ["s.rate / pcs", "s.rate/pcs", "rate / pcs"],
    "gross":         ["gross"],
    "qty_per_ctn":   ["qty per ctn"],
    "qty_ctn":       ["sal qty in ctn"],
    "base_link_doc": ["base link doc", "base link doc  number", "base link doc number"],
    "segment":       ["customer segment name", "customer segment"],
}


def _num(v) -> float:
    s = str(v if v is not None else "").replace(",", "").replace("\xa0", "").strip()
    if not s or s.lower() == "nan":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_sales_excel(data: bytes) -> tuple[list[dict], dict]:
    """Parse the Sales day book Excel. Returns (rows, stats).
    Finds the header row dynamically (the row whose first cells include
    'Date' and 'Voucher'), so leading report banners don't matter."""
    raw = pd.read_excel(io.BytesIO(data), sheet_name=0, header=None)

    header_idx = None
    for i in range(min(15, len(raw))):
        vals = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
        if "date" in vals and any("voucher" in v for v in vals):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Header row not found — is this a Sales day book export?")

    headers = [str(x).strip() for x in raw.iloc[header_idx].tolist()]
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = headers

    # Resolve each of our fields to the actual column name
    def find_col(cands):
        low = {str(c).strip().lower(): c for c in df.columns}
        for cand in cands:
            if cand in low:
                return low[cand]
        # substring fallback
        for cand in cands:
            for k, orig in low.items():
                if cand in k:
                    return orig
        return None

    resolved = {field: find_col(cands) for field, cands in _COLMAP.items()}
    missing = [f for f in ("date", "voucher", "salesman", "item_name", "qty")
               if not resolved[f]]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. "
                         f"Found: {list(df.columns)}")

    rows = []
    for _, r in df.iterrows():
        item = str(r.get(resolved["item_name"], "") or "").replace("\xa0", " ").strip()
        voucher = str(r.get(resolved["voucher"], "") or "").strip()
        if not item or item.lower() == "nan" or not voucher or voucher.lower() == "nan":
            continue  # footer/summary rows

        # Date normalisation → YYYY-MM-DD
        dval = r.get(resolved["date"])
        try:
            d = pd.to_datetime(dval).strftime("%Y-%m-%d")
        except Exception:
            continue

        unit = str(r.get(resolved["unit"], "") or "").replace("\xa0", " ").strip()
        qty = _num(r.get(resolved["qty"]))
        qty_pieces = _num(r.get(resolved["qty_pieces"])) if resolved["qty_pieces"] else 0.0
        qty_per_ctn = _num(r.get(resolved["qty_per_ctn"])) if resolved["qty_per_ctn"] else 0.0
        qty_ctn = _num(r.get(resolved["qty_ctn"])) if resolved["qty_ctn"] else 0.0

        # Fallbacks when helper columns are blank or absent in the export
        up = unit.lstrip(".").lower()
        is_pieces_unit = any(p in up for p in ("pcs", "pc", "piece"))
        if qty_pieces == 0.0 and qty:
            if is_pieces_unit:
                qty_pieces = qty
            elif qty_per_ctn:
                qty_pieces = qty * qty_per_ctn
        # Sal Qty in CTN = Quantity in base unit / Qty Per CTN
        if qty_ctn == 0.0 and qty_pieces and qty_per_ctn:
            qty_ctn = qty_pieces / qty_per_ctn

        rate = _num(r.get(resolved["rate"])) if resolved["rate"] else 0.0
        rate_pcs = _num(r.get(resolved["rate_pcs"])) if resolved["rate_pcs"] else 0.0
        # S.RATE / PCS = Rate / Qty Per CTN (CTN-type rows);
        # for pcs-unit rows the Rate is already per piece
        if rate_pcs == 0.0 and rate:
            if is_pieces_unit:
                rate_pcs = rate
            elif qty_per_ctn:
                rate_pcs = round(rate / qty_per_ctn, 4)

        rows.append({
            "date": d,
            "voucher": voucher,
            "salesman": str(r.get(resolved["salesman"], "") or "").strip() or "UNKNOWN",
            "customer": str(r.get(resolved["customer"], "") or "").strip(),
            "item_name": item,
            "qty": qty,
            "unit": unit,
            "qty_pieces": qty_pieces,
            "qty_ctn": qty_ctn,
            "rate": rate,
            "rate_pcs": rate_pcs,
            "gross": _num(r.get(resolved["gross"])) if resolved["gross"] else 0.0,
            "qty_per_ctn": qty_per_ctn,
            "base_link_doc": str(r.get(resolved["base_link_doc"], "") or "").strip()
                             if resolved["base_link_doc"] else "",
            "segment": str(r.get(resolved["segment"], "") or "").strip()
                       if resolved["segment"] else "",
            "is_foc": 1 if "foc" in unit.lower() else 0,
        })

    if not rows:
        raise ValueError("No data rows found in the file.")

    dates = sorted(r["date"] for r in rows)
    stats = {
        "rows": len(rows),
        "date_from": dates[0],
        "date_to": dates[-1],
        "items": len({r["item_name"] for r in rows}),
        "salesmen": len({r["salesman"] for r in rows}),
        "foc_rows": sum(r["is_foc"] for r in rows),
    }
    return rows, stats


def upsert_sales(conn: sqlite3.Connection, rows: list[dict]) -> dict:
    """Replace-by-(voucher,item) dedupe: delete existing rows for every
    (voucher, item) pair present in the upload, then insert all upload rows.
    Safe for brand-filtered partial files and for Focus corrections."""
    pairs = {(r["voucher"], r["item_name"]) for r in rows}
    deleted = 0
    for v, i in pairs:
        cur = conn.execute(
            "DELETE FROM sales_data WHERE voucher=? AND item_name=?", (v, i))
        deleted += cur.rowcount
    conn.executemany("""
        INSERT INTO sales_data
            (date, voucher, salesman, customer, item_name, qty, unit,
             qty_pieces, qty_ctn, rate, rate_pcs, gross, qty_per_ctn,
             base_link_doc, segment, is_foc)
        VALUES (:date,:voucher,:salesman,:customer,:item_name,:qty,:unit,
                :qty_pieces,:qty_ctn,:rate,:rate_pcs,:gross,:qty_per_ctn,
                :base_link_doc,:segment,:is_foc)
    """, rows)
    conn.commit()
    return {"inserted": len(rows), "replaced": deleted}


# ────────────────────────────────────────────────────────────────────
# Report building
# ────────────────────────────────────────────────────────────────────

def build_report(conn: sqlite3.Connection, group_id: int,
                 date_from: str, date_to: str) -> dict:
    """Salesman x item quantity report for one product group + date range.

    Returns:
    {
      group, date_from, date_to, generated_at,
      salesmen: [names sorted by total pieces desc],
      items: [{
         item, unit_ctn (qty_per_ctn observed),
         per_salesman: {name: {pieces, ctn, foc_pieces}},
         total: {pieces, ctn, foc_pieces},
         target: {qty, unit, achieved_pct} | None,
         salesman_targets: [{salesman, qty, unit, actual, achieved_pct}],
         zero_sellers: [names]
      }],
      grand: {pieces, ctn, foc_pieces},
      ranking: [{salesman, pieces, ctn}]
    }
    """
    g = conn.execute("SELECT name FROM product_groups WHERE id=?",
                     (group_id,)).fetchone()
    if not g:
        raise ValueError("Group not found")
    group_name = g[0]

    item_rows = conn.execute(
        "SELECT item_name FROM group_items WHERE group_id=? ORDER BY item_name",
        (group_id,)).fetchall()
    items = [r[0] for r in item_rows]
    if not items:
        raise ValueError("Group has no items")

    # Salesmen config: names flagged 0 (or unknown, once configured)
    # roll up into one DIRECT SALES bucket
    _cfg = {r[0]: r[1] for r in conn.execute(
        "SELECT name, is_salesman FROM salesmen_config").fetchall()}
    _configured = len(_cfg) > 0

    def _bucket(name: str) -> str:
        if not _configured:
            return name
        return name if _cfg.get(name, 0) else "DIRECT SALES"

    q_marks = ",".join("?" * len(items))
    data = conn.execute(f"""
        SELECT item_name, salesman, is_foc,
               SUM(qty_pieces) AS pieces, SUM(qty_ctn) AS ctn,
               MAX(qty_per_ctn) AS per_ctn
        FROM sales_data
        WHERE item_name IN ({q_marks}) AND date >= ? AND date <= ?
        GROUP BY item_name, salesman, is_foc
    """, (*items, date_from, date_to)).fetchall()
    data = [(it, _bucket(sm), foc, p, ctn, pc)
            for (it, sm, foc, p, ctn, pc) in data]

    targets = conn.execute("""
        SELECT item_name, salesman, target_qty, target_unit
        FROM group_targets WHERE group_id=?
    """, (group_id,)).fetchall()

    # Salesmen = anyone who sold anything in range for these items
    salesmen_set = {r[1] for r in data}
    # plus anyone with a personal target (so their 0 shows)
    for t in targets:
        if t[1]:
            salesmen_set.add(t[1])

    # Accumulate
    acc = {}   # item -> salesman -> {pieces, ctn, foc}
    per_ctn_map = {}
    for item, sman, is_foc, pieces, ctn, per_ctn in data:
        cell = acc.setdefault(item, {}).setdefault(
            sman, {"pieces": 0.0, "ctn": 0.0, "foc_pieces": 0.0})
        if is_foc:
            cell["foc_pieces"] += pieces or 0
        else:
            cell["pieces"] += pieces or 0
            cell["ctn"] += ctn or 0
        if per_ctn:
            per_ctn_map[item] = max(per_ctn_map.get(item, 0), per_ctn)

    # Salesman ranking by total sold pieces
    totals_by_sman = {}
    for item, smap in acc.items():
        for sman, cell in smap.items():
            t = totals_by_sman.setdefault(sman, {"pieces": 0.0, "ctn": 0.0})
            t["pieces"] += cell["pieces"]
            t["ctn"] += cell["ctn"]
    ranking = sorted(
        ({"salesman": s, "pieces": round(v["pieces"], 1),
          "ctn": round(v["ctn"], 2)} for s, v in totals_by_sman.items()),
        key=lambda x: -x["pieces"])
    salesmen = [r["salesman"] for r in ranking]
    for s in sorted(salesmen_set - set(salesmen)):
        salesmen.append(s)  # target-only people with zero sales at the end

    def _ach(actual_pieces, actual_ctn, tqty, tunit):
        actual = actual_ctn if (tunit or "CTN").upper() == "CTN" else actual_pieces
        if not tqty:
            return None
        return round(actual / tqty * 100, 1)

    out_items = []
    grand = {"pieces": 0.0, "ctn": 0.0, "foc_pieces": 0.0}
    for item in items:
        smap = acc.get(item, {})
        per_salesman = {}
        tot = {"pieces": 0.0, "ctn": 0.0, "foc_pieces": 0.0}
        for sman in salesmen:
            cell = smap.get(sman, {"pieces": 0.0, "ctn": 0.0, "foc_pieces": 0.0})
            per_salesman[sman] = {
                "pieces": round(cell["pieces"], 1),
                "ctn": round(cell["ctn"], 2),
                "foc_pieces": round(cell["foc_pieces"], 1),
            }
            for k in tot:
                tot[k] += cell[k]
        for k in grand:
            grand[k] += tot[k]

        # Item-level (overall) target = row with salesman NULL/''
        overall_t = next((t for t in targets
                          if t[0] == item and not t[1]), None)
        target = None
        if overall_t:
            target = {
                "qty": overall_t[2], "unit": (overall_t[3] or "CTN").upper(),
                "achieved_pct": _ach(tot["pieces"], tot["ctn"],
                                     overall_t[2], overall_t[3]),
            }

        sman_targets = []
        for t in targets:
            if t[0] == item and t[1]:
                cell = per_salesman.get(t[1],
                        {"pieces": 0.0, "ctn": 0.0, "foc_pieces": 0.0})
                unit = (t[3] or "CTN").upper()
                actual = cell["ctn"] if unit == "CTN" else cell["pieces"]
                sman_targets.append({
                    "salesman": t[1], "qty": t[2], "unit": unit,
                    "actual": actual,
                    "achieved_pct": _ach(cell["pieces"], cell["ctn"], t[2], t[3]),
                })
        sman_targets.sort(key=lambda x: (x["achieved_pct"] is None,
                                         -(x["achieved_pct"] or 0)))

        zero = [s for s in salesmen
                if per_salesman[s]["pieces"] == 0 and per_salesman[s]["foc_pieces"] == 0]

        out_items.append({
            "item": item,
            "qty_per_ctn": per_ctn_map.get(item, 0),
            "per_salesman": per_salesman,
            "total": {k: round(v, 2) for k, v in tot.items()},
            "target": target,
            "salesman_targets": sman_targets,
            "zero_sellers": zero,
        })

    return {
        "group": group_name,
        "group_id": group_id,
        "date_from": date_from,
        "date_to": date_to,
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "salesmen": salesmen,
        "items": out_items,
        "grand": {k: round(v, 2) for k, v in grand.items()},
        "ranking": ranking,
    }


# ────────────────────────────────────────────────────────────────────
# PDF
# ────────────────────────────────────────────────────────────────────

def report_pdf(report: dict, out_path: str) -> str:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle)

    INK = colors.HexColor("#1a2340")
    ACCENT = colors.HexColor("#0e7d6b")
    SOFT = colors.HexColor("#e9f5f2")
    GREY = colors.HexColor("#5a6379")
    LINE = colors.HexColor("#d5dde5")
    RED = colors.HexColor("#c0392b")
    GREEN = colors.HexColor("#1e8449")

    ss = getSampleStyleSheet()
    H = ParagraphStyle("H", parent=ss["Title"], fontName="Helvetica-Bold",
                       fontSize=17, textColor=INK, spaceAfter=2)
    S = ParagraphStyle("S", parent=ss["Normal"], fontSize=9.5, textColor=GREY,
                       spaceAfter=8)
    H2 = ParagraphStyle("H2", parent=ss["Normal"], fontName="Helvetica-Bold",
                        fontSize=11, textColor=ACCENT, spaceBefore=10,
                        spaceAfter=4)
    CE = ParagraphStyle("CE", parent=ss["Normal"], fontSize=8, leading=10,
                        textColor=INK)
    CEB = ParagraphStyle("CEB", parent=CE, fontName="Helvetica-Bold")

    page = landscape(A4) if len(report["salesmen"]) > 6 else A4
    doc = SimpleDocTemplate(out_path, pagesize=page,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title=f"Sales Report - {report['group']}")
    story = []
    story.append(Paragraph(f"Sales Report — {report['group']}", H))
    story.append(Paragraph(
        f"{report['date_from']} to {report['date_to']} · generated "
        f"{report['generated_at']} · quantities in CTN (pieces below)", S))

    # Matrix table: rows = items, cols = salesmen + total + target + ach
    smen = report["salesmen"]
    has_targets = any(i["target"] for i in report["items"])
    header = ["Item"] + smen + ["TOTAL"]
    if has_targets:
        header += ["Target", "Ach%"]
    data = [[Paragraph(f"<b>{h}</b>",
             ParagraphStyle("hh", parent=CE, textColor=colors.white,
                            fontName="Helvetica-Bold")) for h in header]]

    def cellfmt(c):
        if c["ctn"]:
            s = f"{c['ctn']:g} ctn<br/><font size=6.5 color='#5a6379'>{c['pieces']:g} pcs</font>"
        elif c["pieces"]:
            s = f"{c['pieces']:g} pcs"
        else:
            s = "-"
        if c.get("foc_pieces"):
            s += f"<br/><font size=6.5 color='#b58a00'>FOC {c['foc_pieces']:g}</font>"
        return Paragraph(s, CE)

    for it in report["items"]:
        row = [Paragraph(f"<b>{it['item']}</b>", CE)]
        for s in smen:
            row.append(cellfmt(it["per_salesman"][s]))
        row.append(cellfmt(it["total"]))
        if has_targets:
            if it["target"]:
                t = it["target"]
                row.append(Paragraph(f"{t['qty']:g} {t['unit'].lower()}", CE))
                pct = t["achieved_pct"]
                col = GREEN if (pct or 0) >= 100 else (RED if (pct or 0) < 60 else INK)
                row.append(Paragraph(
                    f"<font color='{col.hexval() if hasattr(col,'hexval') else col}'>"
                    f"<b>{pct if pct is not None else '-'}%</b></font>", CE))
            else:
                row += [Paragraph("-", CE), Paragraph("-", CE)]
        data.append(row)

    n_cols = len(header)
    avail = page[0] - 28*mm
    item_w = min(62*mm, avail * 0.28)
    other_w = (avail - item_w) / (n_cols - 1)
    t = Table(data, colWidths=[item_w] + [other_w] * (n_cols - 1),
              repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
    ]
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), SOFT))
    t.setStyle(TableStyle(style))
    story.append(t)

    # Per-salesman targets
    rows_t = []
    for it in report["items"]:
        for st_ in it["salesman_targets"]:
            pct = st_["achieved_pct"]
            col = "#1e8449" if (pct or 0) >= 100 else ("#c0392b" if (pct or 0) < 60 else "#1a2340")
            rows_t.append([
                Paragraph(it["item"], CE),
                Paragraph(st_["salesman"], CE),
                Paragraph(f"{st_['actual']:g} / {st_['qty']:g} {st_['unit'].lower()}", CE),
                Paragraph(f"<font color='{col}'><b>{pct if pct is not None else '-'}%</b></font>", CE),
            ])
    if rows_t:
        story.append(Paragraph("Personal targets", H2))
        tt = Table([[Paragraph(f"<b>{h}</b>", ParagraphStyle(
                    "hh2", parent=CE, textColor=colors.white))
                    for h in ["Item", "Salesman", "Actual / Target", "Ach%"]]]
                   + rows_t,
                   colWidths=[avail*0.4, avail*0.25, avail*0.22, avail*0.13],
                   repeatRows=1)
        tt.setStyle(TableStyle(style[:7] + [("GRID", (0, 0), (-1, -1), 0.5, LINE),
                    ("BACKGROUND", (0, 0), (-1, 0), ACCENT)]))
        story.append(tt)

    # Zero sellers summary
    zero_lines = []
    for it in report["items"]:
        if it["zero_sellers"]:
            zero_lines.append(f"<b>{it['item']}</b>: "
                              + ", ".join(it["zero_sellers"]))
    if zero_lines:
        story.append(Paragraph("Zero sellers (no sales in range)", H2))
        for z in zero_lines:
            story.append(Paragraph(z, CE))

    # Ranking
    story.append(Paragraph("Salesman ranking (group total)", H2))
    for i, r in enumerate(report["ranking"], 1):
        story.append(Paragraph(
            f"{i}. <b>{r['salesman']}</b> — {r['ctn']:g} ctn "
            f"({r['pieces']:g} pcs)", CE))

    doc.build(story)
    return out_path


# ────────────────────────────────────────────────────────────────────
# Product-list upload (create/extend a group from a CSV or Excel)
# ────────────────────────────────────────────────────────────────────

_HEADER_WORDS = {"product", "products", "product name", "item", "items",
                 "item name", "description", "name", "product list"}


def parse_product_list(data: bytes, filename: str = "") -> list[str]:
    """Parse a one-column product list (CSV or Excel). Tolerates:
    - header variations (PRODUCT / Item Name / etc.) or no header at all
    - BOM, blank lines, duplicate names, extra columns (first column wins)
    Returns the cleaned, de-duplicated product names in file order."""
    name = (filename or "").lower()
    values = []
    if name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(data), sheet_name=0, header=None)
        if df.empty:
            return []
        values = [str(v) for v in df.iloc[:, 0].tolist()]
    else:
        text = data.decode("utf-8-sig", errors="replace")
        for line in text.splitlines():
            # take the first CSV cell only
            cell = line.split(",")[0]
            values.append(cell)

    out, seen = [], set()
    for i, v in enumerate(values):
        v = str(v).replace("\xa0", " ").strip().strip('"').strip()
        if not v or v.lower() == "nan":
            continue
        # Skip a header-looking first row
        if i == 0 and v.lower().strip() in _HEADER_WORDS:
            continue
        key = v.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(v)
    return out


# ────────────────────────────────────────────────────────────────────
# Company-wide sales report (full or single salesman)
# ────────────────────────────────────────────────────────────────────

def build_company_report(conn: sqlite3.Connection, date_from: str,
                         date_to: str, salesman: str = None,
                         products_only: bool = False,
                         group_id: int = None) -> dict:
    """Overall company sales for a date range — all salesmen or one.
    Salesman names are bucketed through salesmen_config (DIRECT SALES)."""
    _cfg = {r[0]: r[1] for r in conn.execute(
        "SELECT name, is_salesman FROM salesmen_config").fetchall()}
    _configured = len(_cfg) > 0

    def _bucket(name: str) -> str:
        if not _configured:
            return name
        return name if _cfg.get(name, 0) else "DIRECT SALES"

    group_name, item_filter, item_params = None, "", []
    if group_id:
        g = conn.execute("SELECT name FROM product_groups WHERE id=?",
                         (group_id,)).fetchone()
        if not g:
            raise ValueError("Group not found")
        group_name = g[0]
        g_items = [r[0] for r in conn.execute(
            "SELECT item_name FROM group_items WHERE group_id=?",
            (group_id,)).fetchall()]
        if not g_items:
            raise ValueError("Group has no items")
        item_filter = f" AND item_name IN ({','.join('?'*len(g_items))})"
        item_params = g_items

    rows = conn.execute(f"""
        SELECT salesman, item_name, customer, is_foc,
               SUM(qty_pieces), SUM(qty_ctn), SUM(gross), COUNT(DISTINCT voucher)
        FROM sales_data
        WHERE date >= ? AND date <= ?{item_filter}
        GROUP BY salesman, item_name, customer, is_foc
    """, (date_from, date_to, *item_params)).fetchall()

    per_sman = {}     # bucket -> {gross, pcs, ctn, foc_pcs, vouchers:set, items:{}, customers:{}}
    for sman, item, cust, foc, pcs, ctn, gross, _v in rows:
        b = _bucket(sman)
        if salesman and b != salesman:
            continue
        s = per_sman.setdefault(b, {"gross": 0.0, "pcs": 0.0, "ctn": 0.0,
                                    "foc_pcs": 0.0, "items": {}, "customers": {}})
        if foc:
            s["foc_pcs"] += pcs or 0
            it = s["items"].setdefault(item, {"pcs": 0.0, "ctn": 0.0,
                                              "gross": 0.0, "foc_pcs": 0.0})
            it["foc_pcs"] += pcs or 0
        else:
            s["gross"] += gross or 0
            s["pcs"] += pcs or 0
            s["ctn"] += ctn or 0
            it = s["items"].setdefault(item, {"pcs": 0.0, "ctn": 0.0,
                                              "gross": 0.0, "foc_pcs": 0.0})
            it["pcs"] += pcs or 0
            it["ctn"] += ctn or 0
            it["gross"] += gross or 0
            cu = s["customers"].setdefault(cust or "-", 0.0)
            s["customers"][cust or "-"] = cu + (gross or 0)

    # voucher counts per bucket (separate pass — DISTINCT within bucket)
    vrows = conn.execute(f"""
        SELECT salesman, COUNT(DISTINCT voucher)
        FROM sales_data WHERE date >= ? AND date <= ?{item_filter}
        GROUP BY salesman
    """, (date_from, date_to, *item_params)).fetchall()
    vch = {}
    for sman, n in vrows:
        b = _bucket(sman)
        vch[b] = vch.get(b, 0) + n

    salesmen_out = []
    for b, s in per_sman.items():
        salesmen_out.append({
            "salesman": b,
            "gross": round(s["gross"], 2),
            "pcs": round(s["pcs"], 1),
            "ctn": round(s["ctn"], 2),
            "foc_pcs": round(s["foc_pcs"], 1),
            "vouchers": vch.get(b, 0),
            "distinct_items": len(s["items"]),
        })
    salesmen_out.sort(key=lambda x: -x["gross"])
    grand_gross = sum(x["gross"] for x in salesmen_out) or 0.0
    for x in salesmen_out:
        x["share_pct"] = round(x["gross"] / grand_gross * 100, 1) if grand_gross else 0

    out = {
        "scope": ("COMPANY (combined — no salesman split)" if products_only
                  else (salesman or "FULL COMPANY"))
                 + (f" · {group_name}" if group_name else ""),
        "no_breakdown": bool(products_only),
        "date_from": date_from,
        "date_to": date_to,
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "grand": {
            "gross": round(grand_gross, 2),
            "pcs": round(sum(x["pcs"] for x in salesmen_out), 1),
            "ctn": round(sum(x["ctn"] for x in salesmen_out), 2),
            "foc_pcs": round(sum(x["foc_pcs"] for x in salesmen_out), 1),
            "vouchers": sum(x["vouchers"] for x in salesmen_out),
        },
        "salesmen": salesmen_out,
    }

    # Item detail: top items overall, or the full item list for one salesman
    items_agg = {}
    cust_agg = {}
    for b, s in per_sman.items():
        for item, it in s["items"].items():
            a = items_agg.setdefault(item, {"pcs": 0.0, "ctn": 0.0,
                                            "gross": 0.0, "foc_pcs": 0.0})
            for k in a:
                a[k] += it[k]
        for cu, g in s["customers"].items():
            cust_agg[cu] = cust_agg.get(cu, 0.0) + g

    items_list = [{"item": k, **{kk: round(vv, 2) for kk, vv in v.items()}}
                  for k, v in items_agg.items()]
    items_list.sort(key=lambda x: -x["gross"])
    full_list = bool(salesman) or products_only
    out["items"] = items_list if full_list else items_list[:20]
    out["items_truncated"] = (not full_list) and len(items_list) > 20
    out["items_total_count"] = len(items_list)

    top_cust = sorted(cust_agg.items(), key=lambda x: -x[1])[:10]
    out["top_customers"] = [{"customer": k, "gross": round(v, 2)}
                            for k, v in top_cust]
    return out


def company_report_pdf(report: dict, out_path: str,
                       show_money: bool = True) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle)

    INK = colors.HexColor("#1a2340")
    ACCENT = colors.HexColor("#155e75")
    SOFT = colors.HexColor("#eef6fa")
    GREY = colors.HexColor("#5a6379")
    LINE = colors.HexColor("#d5dde5")

    ss = getSampleStyleSheet()
    H = ParagraphStyle("H", parent=ss["Title"], fontName="Helvetica-Bold",
                       fontSize=16, textColor=INK, spaceAfter=2)
    S = ParagraphStyle("S", parent=ss["Normal"], fontSize=9.5,
                       textColor=GREY, spaceAfter=8)
    H2 = ParagraphStyle("H2", parent=ss["Normal"], fontName="Helvetica-Bold",
                        fontSize=11, textColor=ACCENT, spaceBefore=10,
                        spaceAfter=4)
    CE = ParagraphStyle("CE", parent=ss["Normal"], fontSize=8.2, leading=10.5,
                        textColor=INK)

    def hdr_cell(t):
        return Paragraph(f"<b>{t}</b>", ParagraphStyle(
            "hh", parent=CE, textColor=colors.white))

    def money(v):
        return f"${v:,.2f}"

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title=f"Sales Report - {report['scope']}")
    avail = A4[0] - 28*mm
    story = []
    story.append(Paragraph(f"Company Sales Report — {report['scope']}", H))
    story.append(Paragraph(
        f"{report['date_from']} to {report['date_to']} · generated "
        f"{report['generated_at']}", S))

    g = report["grand"]
    _cells = []
    if show_money:
        _cells.append(Paragraph(f"<b>{money(g['gross'])}</b><br/>"
                    f"<font size=7.5 color='#5a6379'>GROSS SALES</font>", CE))
    _cells += [
          Paragraph(f"<b>{g['ctn']:,.2f}</b><br/>"
                    f"<font size=7.5 color='#5a6379'>CARTONS</font>", CE),
          Paragraph(f"<b>{g['pcs']:,.0f}</b><br/>"
                    f"<font size=7.5 color='#5a6379'>PIECES</font>", CE),
          Paragraph(f"<b>{g['vouchers']:,}</b><br/>"
                    f"<font size=7.5 color='#5a6379'>INVOICES</font>", CE),
          Paragraph(f"<b>{g['foc_pcs']:,.0f}</b><br/>"
                    f"<font size=7.5 color='#5a6379'>FOC PCS</font>", CE)]
    story.append(Table(
        [_cells],
        colWidths=[avail/len(_cells)]*len(_cells),
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), SOFT),
            ("BOX", (0, 0), (-1, -1), 1, ACCENT),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ])))

    if len(report["salesmen"]) > 1 and not report.get("no_breakdown"):
        story.append(Paragraph("By salesman", H2))
        heads = (["Salesman", "Gross", "Share", "CTN", "Pieces",
                  "FOC pcs", "Invoices", "Items"] if show_money else
                 ["Salesman", "CTN", "Pieces", "FOC pcs", "Invoices", "Items"])
        data = [[hdr_cell(h) for h in heads]]
        for s in report["salesmen"]:
            row = [Paragraph(f"<b>{s['salesman']}</b>", CE)]
            if show_money:
                row += [Paragraph(money(s["gross"]), CE),
                        Paragraph(f"{s['share_pct']}%", CE)]
            row += [Paragraph(f"{s['ctn']:,.2f}", CE),
                    Paragraph(f"{s['pcs']:,.0f}", CE),
                    Paragraph(f"{s['foc_pcs']:,.0f}", CE),
                    Paragraph(f"{s['vouchers']:,}", CE),
                    Paragraph(f"{s['distinct_items']:,}", CE)]
            data.append(row)
        _nc = len(heads) - 1
        w = [avail*0.24] + [avail*0.76/_nc]*_nc
        t = Table(data, colWidths=w, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), INK),
            ("GRID", (0, 0), (-1, -1), 0.5, LINE),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            *[("BACKGROUND", (0, i), (-1, i), SOFT)
              for i in range(2, len(data), 2)]
        ]))
        story.append(t)

    if report.get("no_breakdown") or report["scope"] != "FULL COMPANY":
        title = f"All items sold ({report['items_total_count']:,})"
    else:
        title = f"Top items (of {report['items_total_count']:,})"
    story.append(Paragraph(title, H2))
    iheads = (["Item", "CTN", "Pieces", "FOC pcs", "Gross"] if show_money
              else ["Item", "CTN", "Pieces", "FOC pcs"])
    data = [[hdr_cell(h) for h in iheads]]
    for it in report["items"]:
        row = [Paragraph(it["item"], CE),
               Paragraph(f"{it['ctn']:,.2f}", CE),
               Paragraph(f"{it['pcs']:,.0f}", CE),
               Paragraph(f"{it['foc_pcs']:,.0f}" if it["foc_pcs"] else "-", CE)]
        if show_money:
            row.append(Paragraph(money(it["gross"]), CE))
        data.append(row)
    iw = ([avail*0.48, avail*0.12, avail*0.12, avail*0.12, avail*0.16]
          if show_money else [avail*0.55, avail*0.15, avail*0.15, avail*0.15])
    t = Table(data, colWidths=iw, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        *[("BACKGROUND", (0, i), (-1, i), SOFT)
          for i in range(2, len(data), 2)]
    ]))
    story.append(t)

    if report["top_customers"] and show_money:
        story.append(Paragraph("Top customers", H2))
        for i, cu in enumerate(report["top_customers"], 1):
            story.append(Paragraph(
                f"{i}. <b>{cu['customer']}</b> — {money(cu['gross'])}", CE))

    doc.build(story)
    return out_path


def company_report_xlsx(report: dict, out_path: str,
                        show_money: bool = True) -> str:
    """Company sales report as an Excel workbook: Summary sheet
    (totals + by-salesman) and Items sheet; Top Customers when $ shown."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    INK = "1A2340"
    ACCENT = "155E75"
    SOFT = "EEF6FA"

    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=INK)
    acc_fill = PatternFill("solid", fgColor=ACCENT)
    soft_fill = PatternFill("solid", fgColor=SOFT)
    thin = Border(*[Side(style="thin", color="D5DDE5")]*4)
    money_fmt = '"$"#,##0.00'
    qty_fmt = '#,##0.00'
    int_fmt = '#,##0'

    wb = Workbook()

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Company Sales Report — {report['scope']}"
    ws["A1"].font = Font(bold=True, size=14, color=INK)
    ws["A2"] = (f"{report['date_from']} to {report['date_to']} · "
                f"generated {report['generated_at']}")
    ws["A2"].font = Font(size=9, color="5A6379")

    g = report["grand"]
    labels = ([("Gross Sales", g["gross"], money_fmt)] if show_money else []) + [
        ("Cartons", g["ctn"], qty_fmt),
        ("Pieces", g["pcs"], int_fmt),
        ("Invoices", g["vouchers"], int_fmt),
        ("FOC Pieces", g["foc_pcs"], int_fmt)]
    r = 4
    for i, (lab, val, fmt) in enumerate(labels):
        cell = ws.cell(row=r, column=1 + i, value=lab)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = acc_fill
        cell.border = thin
        v = ws.cell(row=r + 1, column=1 + i, value=val)
        v.number_format = fmt
        v.font = Font(bold=True, size=11)
        v.fill = soft_fill
        v.border = thin

    # By-salesman table
    r = 8
    heads = (["Salesman", "Gross", "Share %", "CTN", "Pieces",
              "FOC pcs", "Invoices", "Items"] if show_money else
             ["Salesman", "CTN", "Pieces", "FOC pcs", "Invoices", "Items"])
    if len(report["salesmen"]) > 1 and not report.get("no_breakdown"):
        ws.cell(row=r, column=1, value="By salesman").font = Font(
            bold=True, size=11, color=ACCENT)
        r += 1
        for ci, h in enumerate(heads, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
        for s in report["salesmen"]:
            r += 1
            vals = [s["salesman"]]
            fmts = [None]
            if show_money:
                vals += [s["gross"], s["share_pct"] / 100]
                fmts += [money_fmt, "0.0%"]
            vals += [s["ctn"], s["pcs"], s["foc_pcs"],
                     s["vouchers"], s["distinct_items"]]
            fmts += [qty_fmt, int_fmt, int_fmt, int_fmt, int_fmt]
            for ci, (v, f) in enumerate(zip(vals, fmts), 1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.border = thin
                if f:
                    cell.number_format = f
    ws.column_dimensions["A"].width = 26
    for ci in range(2, len(heads) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "A3"

    # ── Items sheet ──
    ws2 = wb.create_sheet("Items")
    iheads = (["Item", "CTN", "Pieces", "FOC pcs", "Gross"] if show_money
              else ["Item", "CTN", "Pieces", "FOC pcs"])
    for ci, h in enumerate(iheads, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
    for ri, it in enumerate(report["items"], 2):
        vals = [it["item"], it["ctn"], it["pcs"], it["foc_pcs"]]
        fmts = [None, qty_fmt, int_fmt, int_fmt]
        if show_money:
            vals.append(it["gross"]); fmts.append(money_fmt)
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = thin
            if f:
                cell.number_format = f
    ws2.column_dimensions["A"].width = 52
    for ci in range(2, len(iheads) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 13
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(iheads))}{len(report['items'])+1}"

    # ── Top customers ──
    if report["top_customers"] and show_money:
        ws3 = wb.create_sheet("Top Customers")
        for ci, h in enumerate(["#", "Customer", "Gross"], 1):
            cell = ws3.cell(row=1, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
        for i, cu in enumerate(report["top_customers"], 1):
            ws3.cell(row=i+1, column=1, value=i).border = thin
            ws3.cell(row=i+1, column=2, value=cu["customer"]).border = thin
            cell = ws3.cell(row=i+1, column=3, value=cu["gross"])
            cell.number_format = money_fmt; cell.border = thin
        ws3.column_dimensions["B"].width = 46
        ws3.column_dimensions["C"].width = 15

    wb.save(out_path)
    return out_path


# ────────────────────────────────────────────────────────────────────
# Rank report (leaderboard for a product group, with movement)
# ────────────────────────────────────────────────────────────────────

def _cfg_bucket_fn(conn):
    cfg = {r[0]: r[1] for r in conn.execute(
        "SELECT name, is_salesman FROM salesmen_config").fetchall()}
    configured = len(cfg) > 0

    def bucket(name):
        if not configured:
            return name
        return name if cfg.get(name, 0) else "DIRECT SALES"
    return bucket


def _group_items_or_raise(conn, group_id):
    g = conn.execute("SELECT name FROM product_groups WHERE id=?",
                     (group_id,)).fetchone()
    if not g:
        raise ValueError("Group not found")
    items = [r[0] for r in conn.execute(
        "SELECT item_name FROM group_items WHERE group_id=?",
        (group_id,)).fetchall()]
    if not items:
        raise ValueError("Group has no items")
    return g[0], items


def _rank_rows(conn, items, dfrom, dto, bucket):
    q = ",".join("?" * len(items))
    rows = conn.execute(f"""
        SELECT salesman, item_name,
               SUM(qty_pieces), SUM(qty_ctn), SUM(gross)
        FROM sales_data
        WHERE item_name IN ({q}) AND date >= ? AND date <= ? AND is_foc = 0
        GROUP BY salesman, item_name
    """, (*items, dfrom, dto)).fetchall()
    per_s, per_i = {}, {}
    for sman, item, pcs, ctn, gross in rows:
        b = bucket(sman)
        s = per_s.setdefault(b, {"pcs": 0.0, "ctn": 0.0, "gross": 0.0})
        s["pcs"] += pcs or 0; s["ctn"] += ctn or 0; s["gross"] += gross or 0
        i = per_i.setdefault(item, {"pcs": 0.0, "ctn": 0.0, "gross": 0.0})
        i["pcs"] += pcs or 0; i["ctn"] += ctn or 0; i["gross"] += gross or 0
    return per_s, per_i


def _compare_range(date_from: str, date_to: str, mode: str,
                   custom_from: str = None, custom_to: str = None):
    """Resolve the comparison window.
    mode: '' | 'period' | 'week' | 'month' | 'custom'.
    Returns (from, to) or None."""
    from datetime import date, timedelta
    if not mode:
        return None
    if mode == "custom":
        if not (custom_from and custom_to):
            raise ValueError("Custom comparison needs both dates")
        return (custom_from, custom_to)
    d1 = date.fromisoformat(date_from)
    d2 = date.fromisoformat(date_to)
    if mode == "week":
        return ((d1 - timedelta(days=7)).isoformat(),
                (d2 - timedelta(days=7)).isoformat())
    if mode == "month":
        def back(d):
            y, m = (d.year, d.month - 1) if d.month > 1 else (d.year - 1, 12)
            import calendar
            return date(y, m, min(d.day, calendar.monthrange(y, m)[1]))
        return (back(d1).isoformat(), back(d2).isoformat())
    # 'period': the equal-length window immediately before
    span = (d2 - d1).days
    p2 = d1 - timedelta(days=1)
    p1 = p2 - timedelta(days=span)
    return (p1.isoformat(), p2.isoformat())


def build_rank_report(conn: sqlite3.Connection, group_id: int,
                      date_from: str, date_to: str,
                      compare: str = "", custom_from: str = None,
                      custom_to: str = None) -> dict:
    """compare: '' | 'period' | 'week' | 'month' | 'custom'."""
    group_name, items = _group_items_or_raise(conn, group_id)
    bucket = _cfg_bucket_fn(conn)

    per_s, per_i = _rank_rows(conn, items, date_from, date_to, bucket)

    prev_rank, prev_s = {}, {}
    prev_range = _compare_range(date_from, date_to, compare,
                                custom_from, custom_to)
    if prev_range:
        prev_s, _ = _rank_rows(conn, items, prev_range[0], prev_range[1],
                               bucket)
        prev_order = [s for s, _v in sorted(prev_s.items(),
                                            key=lambda x: -x[1]["pcs"])]
        prev_rank = {s: i + 1 for i, s in enumerate(prev_order)}

    total_pcs = sum(v["pcs"] for v in per_s.values()) or 1
    sm_rank = []
    for i, (s, v) in enumerate(sorted(per_s.items(),
                                      key=lambda x: -x[1]["pcs"]), 1):
        row = {
            "rank": i, "salesman": s,
            "ctn": round(v["ctn"], 2), "pcs": round(v["pcs"], 1),
            "gross": round(v["gross"], 2),
            "share_pct": round(v["pcs"] / total_pcs * 100, 1),
        }
        if prev_range:
            pr = prev_rank.get(s)
            pv = prev_s.get(s)
            row["prev_rank"] = pr
            row["change"] = (pr - i) if pr else None   # + = moved up
            row["prev_ctn"] = round(pv["ctn"], 2) if pv else None
            row["delta_pct"] = (round((v["ctn"] - pv["ctn"]) / pv["ctn"] * 100, 1)
                                if pv and pv["ctn"] else None)
        sm_rank.append(row)

    it_rank = [{"rank": i, "item": k,
                "ctn": round(v["ctn"], 2), "pcs": round(v["pcs"], 1),
                "gross": round(v["gross"], 2)}
               for i, (k, v) in enumerate(
                   sorted(per_i.items(), key=lambda x: -x[1]["pcs"]), 1)]

    return {
        "group": group_name, "group_id": group_id,
        "date_from": date_from, "date_to": date_to,
        "compare": compare or None,
        "prev_from": prev_range[0] if prev_range else None,
        "prev_to": prev_range[1] if prev_range else None,
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "salesman_rank": sm_rank,
        "product_rank": it_rank,
    }


def rank_report_pdf(report: dict, out_path: str) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table,
                                    TableStyle)

    INK = colors.HexColor("#1a2340")
    GOLD = colors.HexColor("#b8860b")
    ACCENT = colors.HexColor("#0e7d6b")
    SOFT = colors.HexColor("#e9f5f2")
    GREY = colors.HexColor("#5a6379")
    LINE = colors.HexColor("#d5dde5")
    GREEN = "#1e8449"; RED = "#c0392b"

    ss = getSampleStyleSheet()
    H = ParagraphStyle("H", parent=ss["Title"], fontName="Helvetica-Bold",
                       fontSize=16, textColor=INK, spaceAfter=2)
    S = ParagraphStyle("S", parent=ss["Normal"], fontSize=9.5,
                       textColor=GREY, spaceAfter=8)
    H2 = ParagraphStyle("H2", parent=ss["Normal"], fontName="Helvetica-Bold",
                        fontSize=11.5, textColor=ACCENT, spaceBefore=12,
                        spaceAfter=5)
    CE = ParagraphStyle("CE", parent=ss["Normal"], fontSize=8.6,
                        leading=11, textColor=INK)

    def hc(t):
        return Paragraph(f"<b>{t}</b>", ParagraphStyle(
            "hh", parent=CE, textColor=colors.white))

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title=f"Rank Report - {report['group']}")
    avail = A4[0] - 28*mm
    story = []
    story.append(Paragraph(f"Rank Report — {report['group']}", H))
    sub = f"{report['date_from']} to {report['date_to']}"
    if report.get("prev_from"):
        labels = {"period": "previous period", "week": "previous week",
                  "month": "previous month", "custom": "custom range"}
        sub += (f" · vs {labels.get(report.get('compare'), 'previous')} "
                f"({report['prev_from']} to {report['prev_to']})")
    sub += f" · generated {report['generated_at']}"
    story.append(Paragraph(sub, S))

    def change_cell(r):
        if r["change"] is None:
            return Paragraph(f"<font color='{GREEN}'><b>NEW</b></font>", CE)
        if r["change"] > 0:
            return Paragraph(
                f"<font color='{GREEN}'><b>+{r['change']}</b></font>", CE)
        if r["change"] < 0:
            return Paragraph(
                f"<font color='{RED}'><b>{r['change']}</b></font>", CE)
        return Paragraph("=", CE)

    has_cmp = bool(report.get("prev_from"))
    story.append(Paragraph("Salesman leaderboard", H2))
    heads = ["#", "Salesman", "CTN", "Pieces", "Share"]
    if has_cmp:
        heads += ["Prev CTN", "Change", "Move"]
    data = [[hc(h) for h in heads]]
    for r in report["salesman_rank"]:
        row = [Paragraph(f"<b>{r['rank']}</b>", CE),
               Paragraph(f"<b>{r['salesman']}</b>", CE),
               Paragraph(f"{r['ctn']:,.2f}", CE),
               Paragraph(f"{r['pcs']:,.0f}", CE),
               Paragraph(f"{r['share_pct']}%", CE)]
        if has_cmp:
            row.append(Paragraph(
                f"{r['prev_ctn']:,.2f}" if r.get("prev_ctn") is not None
                else "-", CE))
            dp = r.get("delta_pct")
            if dp is None:
                row.append(Paragraph("-", CE))
            else:
                col = GREEN if dp >= 0 else RED
                row.append(Paragraph(
                    f"<font color='{col}'><b>{'+' if dp >= 0 else ''}{dp}%</b></font>", CE))
            row.append(change_cell(r))
        data.append(row)
    if has_cmp:
        w = [avail*0.05, avail*0.29, avail*0.12, avail*0.12, avail*0.09,
             avail*0.12, avail*0.11, avail*0.10]
    else:
        w = [avail*0.06, avail*0.46, avail*0.18, avail*0.18, avail*0.12]
    t = Table(data, colWidths=w, repeatRows=1)
    style = [("BACKGROUND", (0, 0), (-1, 0), INK),
             ("GRID", (0, 0), (-1, -1), 0.5, LINE),
             ("VALIGN", (0, 0), (-1, -1), "TOP"),
             ("TOPPADDING", (0, 0), (-1, -1), 4.5),
             ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5)]
    if len(data) > 1:
        style.append(("BACKGROUND", (0, 1), (-1, 1),
                      colors.HexColor("#fff7e0")))   # gold row for #1
    for i in range(3, len(data), 2):
        style.append(("BACKGROUND", (0, i), (-1, i), SOFT))
    t.setStyle(TableStyle(style))
    story.append(t)

    story.append(Paragraph("Product ranking (within group)", H2))
    data2 = [[hc(h) for h in ["#", "Product", "CTN", "Pieces"]]]
    for r in report["product_rank"]:
        data2.append([Paragraph(f"<b>{r['rank']}</b>", CE),
                      Paragraph(r["item"], CE),
                      Paragraph(f"{r['ctn']:,.2f}", CE),
                      Paragraph(f"{r['pcs']:,.0f}", CE)])
    t2 = Table(data2, colWidths=[avail*0.06, avail*0.58, avail*0.18,
                                 avail*0.18], repeatRows=1)
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        *[("BACKGROUND", (0, i), (-1, i), SOFT)
          for i in range(2, len(data2), 2)]
    ]))
    story.append(t2)
    doc.build(story)
    return out_path


# ────────────────────────────────────────────────────────────────────
# Daily sales report (by value)
# ────────────────────────────────────────────────────────────────────

def _daily_days(conn, date_from, date_to, salesman, bucket):
    rows = conn.execute("""
        SELECT date, salesman,
               SUM(CASE WHEN is_foc=0 THEN gross ELSE 0 END),
               SUM(CASE WHEN is_foc=0 THEN qty_ctn ELSE 0 END),
               SUM(CASE WHEN is_foc=0 THEN qty_pieces ELSE 0 END),
               COUNT(DISTINCT voucher)
        FROM sales_data
        WHERE date >= ? AND date <= ?
        GROUP BY date, salesman
        ORDER BY date
    """, (date_from, date_to)).fetchall()
    days = {}
    for d, sman, gross, ctn, pcs, vch in rows:
        if salesman and bucket(sman) != salesman:
            continue
        day = days.setdefault(d, {"gross": 0.0, "ctn": 0.0, "pcs": 0.0,
                                  "vouchers": 0})
        day["gross"] += gross or 0
        day["ctn"] += ctn or 0
        day["pcs"] += pcs or 0
        day["vouchers"] += vch or 0
    return days


def build_daily_report(conn: sqlite3.Connection, date_from: str,
                       date_to: str, salesman: str = None,
                       compare: str = "", custom_from: str = None,
                       custom_to: str = None) -> dict:
    """compare: '' | 'period' | 'week' | 'month' | 'custom' — aligns each
    day with the corresponding day of the comparison window by offset."""
    bucket = _cfg_bucket_fn(conn)
    days = _daily_days(conn, date_from, date_to, salesman, bucket)

    prev_range = _compare_range(date_from, date_to, compare,
                                custom_from, custom_to)
    prev_days = (_daily_days(conn, prev_range[0], prev_range[1],
                             salesman, bucket) if prev_range else {})

    from datetime import date as _date, timedelta as _td
    out_days = []
    for d, v in sorted(days.items()):
        row = {"date": d,
               "gross": round(v["gross"], 2),
               "ctn": round(v["ctn"], 2),
               "pcs": round(v["pcs"], 1),
               "vouchers": v["vouchers"]}
        if prev_range:
            offset = (_date.fromisoformat(d)
                      - _date.fromisoformat(date_from)).days
            pd = (_date.fromisoformat(prev_range[0])
                  + _td(days=offset)).isoformat()
            pv = prev_days.get(pd)
            row["prev_date"] = pd
            row["prev_gross"] = round(pv["gross"], 2) if pv else None
            row["change_pct"] = (round((v["gross"] - pv["gross"])
                                       / pv["gross"] * 100, 1)
                                 if pv and pv["gross"] else None)
        out_days.append(row)
    n = len(out_days) or 1
    tot = {
        "gross": round(sum(x["gross"] for x in out_days), 2),
        "ctn": round(sum(x["ctn"] for x in out_days), 2),
        "pcs": round(sum(x["pcs"] for x in out_days), 1),
        "vouchers": sum(x["vouchers"] for x in out_days),
    }
    best = max(out_days, key=lambda x: x["gross"]) if out_days else None
    prev_total = None
    total_change_pct = None
    if prev_range:
        prev_total = round(sum(v["gross"] for v in prev_days.values()), 2)
        if prev_total:
            total_change_pct = round(
                (tot["gross"] - prev_total) / prev_total * 100, 1)
    return {
        "scope": salesman or "FULL COMPANY",
        "date_from": date_from, "date_to": date_to,
        "compare": compare or None,
        "prev_from": prev_range[0] if prev_range else None,
        "prev_to": prev_range[1] if prev_range else None,
        "prev_total": prev_total,
        "total_change_pct": total_change_pct,
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "days": out_days,
        "total": tot,
        "avg_per_day": round(tot["gross"] / n, 2),
        "best_day": best,
    }


def daily_report_pdf(report: dict, out_path: str) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table,
                                    TableStyle)

    INK = colors.HexColor("#1a2340")
    ACCENT = colors.HexColor("#155e75")
    SOFT = colors.HexColor("#eef6fa")
    GREY = colors.HexColor("#5a6379")
    LINE = colors.HexColor("#d5dde5")

    ss = getSampleStyleSheet()
    H = ParagraphStyle("H", parent=ss["Title"], fontName="Helvetica-Bold",
                       fontSize=16, textColor=INK, spaceAfter=2)
    S = ParagraphStyle("S", parent=ss["Normal"], fontSize=9.5,
                       textColor=GREY, spaceAfter=8)
    CE = ParagraphStyle("CE", parent=ss["Normal"], fontSize=8.6,
                        leading=11, textColor=INK)

    def hc(t):
        return Paragraph(f"<b>{t}</b>", ParagraphStyle(
            "hh", parent=CE, textColor=colors.white))

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title=f"Daily Sales - {report['scope']}")
    avail = A4[0] - 28*mm
    story = []
    story.append(Paragraph(f"Daily Sales by Value — {report['scope']}", H))
    sub = f"{report['date_from']} to {report['date_to']}"
    if report.get("prev_from"):
        labels = {"period": "previous period", "week": "previous week",
                  "month": "previous month", "custom": "custom range"}
        sub += (f" · vs {labels.get(report.get('compare'), 'previous')} "
                f"({report['prev_from']} to {report['prev_to']})")
    sub += f" · generated {report['generated_at']}"
    story.append(Paragraph(sub, S))

    t0 = report["total"]
    best = report["best_day"]
    strip = [[Paragraph(f"<b>${t0['gross']:,.2f}</b><br/>"
                        f"<font size=7.5 color='#5a6379'>TOTAL SALES</font>", CE),
              Paragraph(f"<b>${report['avg_per_day']:,.2f}</b><br/>"
                        f"<font size=7.5 color='#5a6379'>AVG / DAY</font>", CE),
              Paragraph(f"<b>{t0['vouchers']:,}</b><br/>"
                        f"<font size=7.5 color='#5a6379'>INVOICES</font>", CE),
              Paragraph((f"<b>{best['date']}</b><br/>"
                         f"<font size=7.5 color='#5a6379'>BEST DAY "
                         f"(${best['gross']:,.0f})</font>") if best else "-",
                        CE)]]
    if report.get("prev_total") is not None:
        tcp = report.get("total_change_pct")
        col = "#1e8449" if (tcp or 0) >= 0 else "#c0392b"
        strip[0].append(Paragraph(
            f"<b><font color='{col}'>"
            f"{('+' if (tcp or 0) >= 0 else '')}{tcp if tcp is not None else '-'}%"
            f"</font></b><br/><font size=7.5 color='#5a6379'>VS PREV "
            f"(${report['prev_total']:,.0f})</font>", CE))
    st = Table(strip, colWidths=[avail/len(strip[0])]*len(strip[0]))
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), SOFT),
        ("BOX", (0, 0), (-1, -1), 1, ACCENT),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10)]))
    story.append(st)
    story.append(Paragraph(" ", S))

    has_cmp = bool(report.get("prev_from"))
    max_g = max((d["gross"] for d in report["days"]), default=1) or 1
    heads = ["Date", "Sales $", "", "CTN", "Pieces", "Invoices"]
    if has_cmp:
        heads += ["Prev $", "Change"]
    data = [[hc(h) for h in heads]]
    for d in report["days"]:
        bar_w = max(2, int(d["gross"] / max_g * 100))
        bar = (f"<font color='#0e7d6b'>{'█' * max(1, bar_w // 7)}</font>")
        row = [Paragraph(d["date"], CE),
               Paragraph(f"<b>${d['gross']:,.2f}</b>", CE),
               Paragraph(bar, CE),
               Paragraph(f"{d['ctn']:,.1f}", CE),
               Paragraph(f"{d['pcs']:,.0f}", CE),
               Paragraph(f"{d['vouchers']:,}", CE)]
        if has_cmp:
            row.append(Paragraph(
                f"${d['prev_gross']:,.2f}" if d.get("prev_gross") is not None
                else "-", CE))
            cp = d.get("change_pct")
            if cp is None:
                row.append(Paragraph("-", CE))
            else:
                col = "#1e8449" if cp >= 0 else "#c0392b"
                row.append(Paragraph(
                    f"<font color='{col}'><b>{'+' if cp >= 0 else ''}{cp}%</b></font>",
                    CE))
        data.append(row)
    trow = [Paragraph("<b>TOTAL</b>", CE),
            Paragraph(f"<b>${t0['gross']:,.2f}</b>", CE),
            Paragraph("", CE),
            Paragraph(f"<b>{t0['ctn']:,.1f}</b>", CE),
            Paragraph(f"<b>{t0['pcs']:,.0f}</b>", CE),
            Paragraph(f"<b>{t0['vouchers']:,}</b>", CE)]
    if has_cmp:
        trow.append(Paragraph(
            f"<b>${report['prev_total']:,.2f}</b>"
            if report.get("prev_total") is not None else "-", CE))
        tcp = report.get("total_change_pct")
        col = "#1e8449" if (tcp or 0) >= 0 else "#c0392b"
        trow.append(Paragraph(
            f"<font color='{col}'><b>{('+' if (tcp or 0) >= 0 else '')}{tcp if tcp is not None else '-'}%</b></font>",
            CE))
    data.append(trow)
    if has_cmp:
        cw = [avail*0.11, avail*0.14, avail*0.19, avail*0.10, avail*0.11,
              avail*0.10, avail*0.13, avail*0.12]
    else:
        cw = [avail*0.14, avail*0.18, avail*0.26, avail*0.14, avail*0.14,
              avail*0.14]
    t = Table(data, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("BACKGROUND", (0, len(data)-1), (-1, len(data)-1), SOFT),
        ("LINEABOVE", (0, len(data)-1), (-1, len(data)-1), 1, ACCENT),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        *[("BACKGROUND", (0, i), (-1, i), SOFT)
          for i in range(2, len(data)-1, 2)]
    ]))
    story.append(t)
    doc.build(story)
    return out_path


def daily_report_xlsx(report: dict, out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1A2340")
    soft = PatternFill("solid", fgColor="EEF6FA")
    thin = Border(*[Side(style="thin", color="D5DDE5")]*4)
    money = '"$"#,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Sales"
    ws["A1"] = f"Daily Sales by Value — {report['scope']}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"{report['date_from']} to {report['date_to']}"
    ws["A2"].font = Font(size=9, color="5A6379")

    has_cmp = bool(report.get("prev_from"))
    heads = ["Date", "Sales $", "CTN", "Pieces", "Invoices"]
    if has_cmp:
        heads += ["Prev $", "Change %"]
        ws["A3"] = (f"Compared with {report['prev_from']} to "
                    f"{report['prev_to']}")
        ws["A3"].font = Font(size=9, color="5A6379")
    for ci, h in enumerate(heads, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
    r = 4
    for d in report["days"]:
        r += 1
        vals = [d["date"], d["gross"], d["ctn"], d["pcs"], d["vouchers"]]
        fmts = [None, money, '#,##0.00', '#,##0', '#,##0']
        if has_cmp:
            vals += [d.get("prev_gross"),
                     (d["change_pct"] / 100) if d.get("change_pct") is not None
                     else None]
            fmts += [money, '+0.0%;-0.0%']
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = thin
            if f and v is not None:
                cell.number_format = f
    r += 1
    t0 = report["total"]
    tvals = ["TOTAL", t0["gross"], t0["ctn"], t0["pcs"], t0["vouchers"]]
    tfmts = [None, money, '#,##0.00', '#,##0', '#,##0']
    if has_cmp:
        tvals += [report.get("prev_total"),
                  (report["total_change_pct"] / 100)
                  if report.get("total_change_pct") is not None else None]
        tfmts += [money, '+0.0%;-0.0%']
    for ci, (v, f) in enumerate(zip(tvals, tfmts), 1):
        cell = ws.cell(row=r, column=ci, value=v)
        cell.font = Font(bold=True); cell.fill = soft; cell.border = thin
        if f and v is not None:
            cell.number_format = f
    ws.column_dimensions["A"].width = 14
    for ci in range(2, len(heads) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = "A5"
    wb.save(out_path)
    return out_path


# ────────────────────────────────────────────────────────────────────
# Customer-wise sales comparison
# ────────────────────────────────────────────────────────────────────

def _customer_totals(conn, dfrom, dto, salesman, bucket):
    rows = conn.execute("""
        SELECT customer, salesman,
               SUM(CASE WHEN is_foc=0 THEN gross ELSE 0 END),
               SUM(CASE WHEN is_foc=0 THEN qty_ctn ELSE 0 END),
               COUNT(DISTINCT voucher)
        FROM sales_data
        WHERE date >= ? AND date <= ?
        GROUP BY customer, salesman
    """, (dfrom, dto)).fetchall()
    out = {}
    for cust, sman, gross, ctn, vch in rows:
        if salesman and bucket(sman) != salesman:
            continue
        cu = out.setdefault(cust or "-", {"gross": 0.0, "ctn": 0.0,
                                          "vouchers": 0})
        cu["gross"] += gross or 0
        cu["ctn"] += ctn or 0
        cu["vouchers"] += vch or 0
    return out


def build_customer_compare(conn: sqlite3.Connection, date_from: str,
                           date_to: str, salesman: str = None,
                           compare: str = "period", custom_from: str = None,
                           custom_to: str = None, limit: int = 200) -> dict:
    bucket = _cfg_bucket_fn(conn)
    prev_range = _compare_range(date_from, date_to, compare or "period",
                                custom_from, custom_to)
    a = _customer_totals(conn, date_from, date_to, salesman, bucket)
    b = _customer_totals(conn, prev_range[0], prev_range[1],
                         salesman, bucket)

    all_customers = set(a) | set(b)
    rows = []
    for cust in all_customers:
        ga = round(a.get(cust, {}).get("gross", 0.0), 2)
        gb = round(b.get(cust, {}).get("gross", 0.0), 2)
        if ga == 0 and gb == 0:
            continue
        rows.append({
            "customer": cust,
            "gross_a": ga,
            "gross_b": gb,
            "ctn_a": round(a.get(cust, {}).get("ctn", 0.0), 2),
            "vouchers_a": a.get(cust, {}).get("vouchers", 0),
            "diff": round(ga - gb, 2),
            "change_pct": (round((ga - gb) / gb * 100, 1) if gb else None),
            "status": ("NEW" if gb == 0 else ("LOST" if ga == 0 else "")),
        })
    rows.sort(key=lambda x: -x["gross_a"])

    tot_a = round(sum(r["gross_a"] for r in rows), 2)
    tot_b = round(sum(r["gross_b"] for r in rows), 2)
    return {
        "scope": salesman or "FULL COMPANY",
        "date_from": date_from, "date_to": date_to,
        "compare": compare or "period",
        "prev_from": prev_range[0], "prev_to": prev_range[1],
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "customers": rows[:limit],
        "customer_count": len(rows),
        "truncated": len(rows) > limit,
        "total_a": tot_a, "total_b": tot_b,
        "total_change_pct": (round((tot_a - tot_b) / tot_b * 100, 1)
                             if tot_b else None),
        "new_count": sum(1 for r in rows if r["status"] == "NEW"),
        "lost_count": sum(1 for r in rows if r["status"] == "LOST"),
    }


def customer_compare_pdf(report: dict, out_path: str) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table,
                                    TableStyle)
    INK = colors.HexColor("#1a2340")
    ACCENT = colors.HexColor("#155e75")
    SOFT = colors.HexColor("#eef6fa")
    GREY = colors.HexColor("#5a6379")
    LINE = colors.HexColor("#d5dde5")
    ss = getSampleStyleSheet()
    H = ParagraphStyle("H", parent=ss["Title"], fontName="Helvetica-Bold",
                       fontSize=15, textColor=INK, spaceAfter=2)
    S = ParagraphStyle("S", parent=ss["Normal"], fontSize=9,
                       textColor=GREY, spaceAfter=8)
    CE = ParagraphStyle("CE", parent=ss["Normal"], fontSize=7.8,
                        leading=10, textColor=INK)

    def hc(t):
        return Paragraph(f"<b>{t}</b>", ParagraphStyle(
            "hh", parent=CE, textColor=colors.white))

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title=f"Customer Comparison - {report['scope']}")
    avail = A4[0] - 24*mm
    story = [Paragraph(f"Customer Sales Comparison — {report['scope']}", H),
             Paragraph(
        f"A: {report['date_from']} to {report['date_to']} · "
        f"B: {report['prev_from']} to {report['prev_to']} · "
        f"{report['customer_count']:,} customers · "
        f"NEW {report['new_count']} · LOST {report['lost_count']} · "
        f"generated {report['generated_at']}", S)]

    tcp = report.get("total_change_pct")
    col = "#1e8449" if (tcp or 0) >= 0 else "#c0392b"
    story.append(Paragraph(
        f"Total A: <b>${report['total_a']:,.2f}</b> · "
        f"Total B: <b>${report['total_b']:,.2f}</b> · Change: "
        f"<font color='{col}'><b>"
        f"{('+' if (tcp or 0) >= 0 else '')}{tcp if tcp is not None else '-'}%"
        f"</b></font>", S))

    data = [[hc(h) for h in ["Customer", "Sales A", "Sales B",
                             "Diff", "Change", ""]]]
    for r in report["customers"]:
        cp = r["change_pct"]
        if cp is None:
            chg = Paragraph("-", CE)
        else:
            ccol = "#1e8449" if cp >= 0 else "#c0392b"
            chg = Paragraph(
                f"<font color='{ccol}'><b>{'+' if cp >= 0 else ''}{cp}%</b></font>", CE)
        badge = ""
        if r["status"] == "NEW":
            badge = "<font color='#1e8449'><b>NEW</b></font>"
        elif r["status"] == "LOST":
            badge = "<font color='#c0392b'><b>LOST</b></font>"
        data.append([Paragraph(r["customer"], CE),
                     Paragraph(f"${r['gross_a']:,.2f}", CE),
                     Paragraph(f"${r['gross_b']:,.2f}", CE),
                     Paragraph(f"{'+' if r['diff'] >= 0 else ''}"
                               f"${r['diff']:,.2f}", CE),
                     chg, Paragraph(badge, CE)])
    t = Table(data, colWidths=[avail*0.38, avail*0.15, avail*0.15,
                               avail*0.14, avail*0.10, avail*0.08],
              repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("GRID", (0, 0), (-1, -1), 0.4, LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        *[("BACKGROUND", (0, i), (-1, i), SOFT)
          for i in range(2, len(data), 2)]
    ]))
    story.append(t)
    doc.build(story)
    return out_path


def customer_compare_xlsx(report: dict, out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1A2340")
    thin = Border(*[Side(style="thin", color="D5DDE5")]*4)
    money = '"$"#,##0.00'

    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Comparison"
    ws["A1"] = f"Customer Sales Comparison — {report['scope']}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"A: {report['date_from']} to {report['date_to']}   |   "
                f"B: {report['prev_from']} to {report['prev_to']}")
    ws["A2"].font = Font(size=9, color="5A6379")
    heads = ["Customer", "Sales A", "Sales B", "Diff", "Change %",
             "CTN (A)", "Invoices (A)", "Status"]
    for ci, h in enumerate(heads, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
    r = 4
    for d in report["customers"]:
        r += 1
        vals = [d["customer"], d["gross_a"], d["gross_b"], d["diff"],
                (d["change_pct"] / 100) if d["change_pct"] is not None
                else None,
                d["ctn_a"], d["vouchers_a"], d["status"]]
        fmts = [None, money, money, money, '+0.0%;-0.0%',
                '#,##0.00', '#,##0', None]
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = thin
            if f and v is not None:
                cell.number_format = f
    ws.column_dimensions["A"].width = 46
    for ci in range(2, len(heads) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(heads))}{r}"
    wb.save(out_path)
    return out_path


# ────────────────────────────────────────────────────────────────────
# Trend report (monthly / weekly buckets)
# ────────────────────────────────────────────────────────────────────

def build_trend(conn: sqlite3.Connection, date_from: str, date_to: str,
                granularity: str = "month", salesman: str = None,
                customer: str = None) -> dict:
    from datetime import date as _date
    bucket_fn = _cfg_bucket_fn(conn)
    where, params = "WHERE date >= ? AND date <= ?", [date_from, date_to]
    if customer:
        where += " AND customer = ?"
        params.append(customer)
    rows = conn.execute(f"""
        SELECT date, salesman,
               SUM(CASE WHEN is_foc=0 THEN gross ELSE 0 END),
               SUM(CASE WHEN is_foc=0 THEN qty_ctn ELSE 0 END),
               SUM(CASE WHEN is_foc=0 THEN qty_pieces ELSE 0 END),
               COUNT(DISTINCT voucher)
        FROM sales_data {where}
        GROUP BY date, salesman
    """, params).fetchall()

    buckets = {}
    for d, sman, gross, ctn, pcs, vch in rows:
        if salesman and bucket_fn(sman) != salesman:
            continue
        dt = _date.fromisoformat(d)
        if granularity == "week":
            iso = dt.isocalendar()
            key = f"{iso[0]}-W{iso[1]:02d}"
        else:
            key = d[:7]
        b = buckets.setdefault(key, {"gross": 0.0, "ctn": 0.0, "pcs": 0.0,
                                     "vouchers": 0, "days": set()})
        b["gross"] += gross or 0
        b["ctn"] += ctn or 0
        b["pcs"] += pcs or 0
        b["vouchers"] += vch or 0
        b["days"].add(d)

    out = []
    prev_g = None
    for key in sorted(buckets):
        b = buckets[key]
        g = round(b["gross"], 2)
        out.append({
            "bucket": key,
            "gross": g,
            "ctn": round(b["ctn"], 2),
            "pcs": round(b["pcs"], 1),
            "vouchers": b["vouchers"],
            "active_days": len(b["days"]),
            "change_pct": (round((g - prev_g) / prev_g * 100, 1)
                           if prev_g else None),
        })
        prev_g = g

    tot = round(sum(x["gross"] for x in out), 2)
    return {
        "scope": (customer or salesman or "FULL COMPANY")
                 + ((" · " + salesman) if (customer and salesman) else ""),
        "granularity": granularity,
        "date_from": date_from, "date_to": date_to,
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "buckets": out,
        "total": tot,
    }


def trend_pdf(report: dict, out_path: str) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table,
                                    TableStyle)
    INK = colors.HexColor("#1a2340")
    ACCENT = colors.HexColor("#0e7d6b")
    SOFT = colors.HexColor("#e9f5f2")
    GREY = colors.HexColor("#5a6379")
    LINE = colors.HexColor("#d5dde5")
    ss = getSampleStyleSheet()
    H = ParagraphStyle("H", parent=ss["Title"], fontName="Helvetica-Bold",
                       fontSize=15, textColor=INK, spaceAfter=2)
    S = ParagraphStyle("S", parent=ss["Normal"], fontSize=9,
                       textColor=GREY, spaceAfter=8)
    CE = ParagraphStyle("CE", parent=ss["Normal"], fontSize=8.6,
                        leading=11, textColor=INK)

    def hc(t):
        return Paragraph(f"<b>{t}</b>", ParagraphStyle(
            "hh", parent=CE, textColor=colors.white))

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=12*mm, bottomMargin=12*mm,
                            title=f"Sales Trend - {report['scope']}")
    avail = A4[0] - 28*mm
    gran = "Monthly" if report["granularity"] == "month" else "Weekly"
    story = [Paragraph(f"{gran} Sales Trend — {report['scope']}", H),
             Paragraph(f"{report['date_from']} to {report['date_to']} · "
                       f"total ${report['total']:,.2f} · generated "
                       f"{report['generated_at']}", S)]

    max_g = max((b["gross"] for b in report["buckets"]), default=1) or 1
    data = [[hc(h) for h in [gran[:-2] if gran.endswith("ly") else gran,
                             "Sales $", "", "Change", "CTN", "Invoices",
                             "Days"]]]
    for b in report["buckets"]:
        bar_n = max(1, int(b["gross"] / max_g * 14))
        bar = f"<font color='#0e7d6b'>{'█' * bar_n}</font>"
        cp = b["change_pct"]
        if cp is None:
            chg = Paragraph("-", CE)
        else:
            col = "#1e8449" if cp >= 0 else "#c0392b"
            chg = Paragraph(
                f"<font color='{col}'><b>{'+' if cp >= 0 else ''}{cp}%</b></font>", CE)
        data.append([Paragraph(f"<b>{b['bucket']}</b>", CE),
                     Paragraph(f"<b>${b['gross']:,.2f}</b>", CE),
                     Paragraph(bar, CE),
                     chg,
                     Paragraph(f"{b['ctn']:,.1f}", CE),
                     Paragraph(f"{b['vouchers']:,}", CE),
                     Paragraph(f"{b['active_days']}", CE)])
    t = Table(data, colWidths=[avail*0.13, avail*0.17, avail*0.28,
                               avail*0.12, avail*0.12, avail*0.10,
                               avail*0.08], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("GRID", (0, 0), (-1, -1), 0.5, LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        *[("BACKGROUND", (0, i), (-1, i), SOFT)
          for i in range(2, len(data), 2)]
    ]))
    story.append(t)
    doc.build(story)
    return out_path


def trend_xlsx(report: dict, out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1A2340")
    thin = Border(*[Side(style="thin", color="D5DDE5")]*4)
    money = '"$"#,##0.00'
    wb = Workbook()
    ws = wb.active
    gran = "Monthly" if report["granularity"] == "month" else "Weekly"
    ws.title = f"{gran} Trend"
    ws["A1"] = f"{gran} Sales Trend — {report['scope']}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"{report['date_from']} to {report['date_to']}"
    ws["A2"].font = Font(size=9, color="5A6379")
    heads = ["Period", "Sales $", "Change %", "CTN", "Pieces", "Invoices",
             "Active days"]
    for ci, h in enumerate(heads, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
    r = 4
    for b in report["buckets"]:
        r += 1
        vals = [b["bucket"], b["gross"],
                (b["change_pct"] / 100) if b["change_pct"] is not None
                else None,
                b["ctn"], b["pcs"], b["vouchers"], b["active_days"]]
        fmts = [None, money, '+0.0%;-0.0%', '#,##0.00', '#,##0',
                '#,##0', '#,##0']
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = thin
            if f and v is not None:
                cell.number_format = f
    ws.column_dimensions["A"].width = 14
    for ci in range(2, len(heads) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "A5"
    wb.save(out_path)
    return out_path


# ────────────────────────────────────────────────────────────────────
# Customer performance (monthly matrix + not-ordered list)
# ────────────────────────────────────────────────────────────────────

def build_customer_performance(conn: sqlite3.Connection, month: str,
                               months_history: int = 3,
                               salesman: str = None) -> dict:
    """month: 'YYYY-MM' (the selected month). Shows each customer's totals
    for the previous `months_history` months, the selected month so far,
    and the difference vs last month — computed fairly against the same
    number of days when the selected month is incomplete."""
    import calendar
    from datetime import date as _date
    bucket = _cfg_bucket_fn(conn)

    y, m = int(month[:4]), int(month[5:7])

    def month_range(yy, mm):
        last = calendar.monthrange(yy, mm)[1]
        return (f"{yy:04d}-{mm:02d}-01", f"{yy:04d}-{mm:02d}-{last:02d}")

    def back(yy, mm, k):
        mm -= k
        while mm < 1:
            mm += 12
            yy -= 1
        return yy, mm

    # Is the selected month partial (data hasn't reached month end)?
    sel_from, sel_to = month_range(y, m)
    max_d = conn.execute(
        "SELECT MAX(date) FROM sales_data WHERE date >= ? AND date <= ?",
        (sel_from, sel_to)).fetchone()[0]
    days_elapsed = int(max_d[8:10]) if max_d else 0
    month_days = calendar.monthrange(y, m)[1]
    partial = bool(max_d) and days_elapsed < month_days
    if partial:
        sel_to_eff = f"{y:04d}-{m:02d}-{days_elapsed:02d}"
    else:
        sel_to_eff = sel_to

    # Month windows: oldest → last month, then selected
    hist = []
    for k in range(months_history, 0, -1):
        hy, hm = back(y, m, k)
        hist.append({"key": f"{hy:04d}-{hm:02d}",
                     "range": month_range(hy, hm)})

    # Last-month same-days window for fair comparison
    ly, lm = back(y, m, 1)
    lm_last = calendar.monthrange(ly, lm)[1]
    lm_same_to = f"{ly:04d}-{lm:02d}-{min(days_elapsed or lm_last, lm_last):02d}"
    lm_same_range = (f"{ly:04d}-{lm:02d}-01", lm_same_to)

    def totals(rng):
        return _customer_totals(conn, rng[0], rng[1], salesman, bucket)

    hist_totals = [totals(h["range"]) for h in hist]
    sel_totals = totals((sel_from, sel_to_eff))
    lm_same_totals = totals(lm_same_range) if partial else None

    all_customers = set(sel_totals)
    for ht in hist_totals:
        all_customers |= set(ht)

    lm_key_totals = hist_totals[-1] if hist_totals else {}
    rows = []
    for cust in all_customers:
        hvals = [round(ht.get(cust, {}).get("gross", 0.0), 2)
                 for ht in hist_totals]
        this_g = round(sel_totals.get(cust, {}).get("gross", 0.0), 2)
        lm_full = hvals[-1] if hvals else 0.0
        base = (round(lm_same_totals.get(cust, {}).get("gross", 0.0), 2)
                if partial and lm_same_totals is not None else lm_full)
        if this_g == 0 and all(v == 0 for v in hvals):
            continue
        rows.append({
            "customer": cust,
            "history": hvals,                    # oldest → last month
            "this_month": this_g,
            "lm_same_days": base if partial else None,
            "diff": round(this_g - base, 2),
            "change_pct": (round((this_g - base) / base * 100, 1)
                           if base else None),
            "not_ordered": this_g == 0 and any(v > 0 for v in hvals),
        })
    rows.sort(key=lambda r: -(r["history"][-1] if r["history"] else 0))

    # Enrich from the customers master when present (segment / area),
    # and find master customers with no sales at all in the window.
    master = {r[0].strip().lower(): {"segment": r[1] or "", "area": r[2] or "",
                                     "mobile": r[3] or ""}
              for r in conn.execute(
                  "SELECT name, segment, area, mobile "
                  "FROM customers_master").fetchall()}
    never_ordered = []
    if master:
        for r in rows:
            m = master.get(r["customer"].strip().lower())
            r["segment"] = m["segment"] if m else ""
            r["area"] = m["area"] if m else ""
        seen = {r["customer"].strip().lower() for r in rows}
        for key, m in master.items():
            if key not in seen:
                never_ordered.append(key)

    not_ordered = [r for r in rows if r["not_ordered"]]
    tot_hist = [round(sum(r["history"][i] for r in rows), 2)
                for i in range(len(hist))]
    return {
        "scope": salesman or "FULL COMPANY",
        "month": month,
        "partial": partial,
        "days_elapsed": days_elapsed if partial else month_days,
        "month_days": month_days,
        "hist_keys": [h["key"] for h in hist],
        "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        "customers": rows,
        "customer_count": len(rows),
        "totals": {
            "history": tot_hist,
            "this_month": round(sum(r["this_month"] for r in rows), 2),
            "lm_same_days": (round(sum(r["lm_same_days"] or 0 for r in rows), 2)
                             if partial else None),
        },
        "has_master": bool(master),
        "never_ordered_count": len(never_ordered) if master else None,
        "not_ordered_count": len(not_ordered),
        "not_ordered_top": [{"customer": r["customer"],
                             "last_month": r["history"][-1]}
                            for r in not_ordered[:25]],
    }


def customer_performance_pdf(report: dict, out_path: str,
                             max_rows: int = 200) -> str:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table,
                                    TableStyle)
    INK = colors.HexColor("#1a2340")
    SOFT = colors.HexColor("#eef6fa")
    GREY = colors.HexColor("#5a6379")
    LINE = colors.HexColor("#d5dde5")
    ss = getSampleStyleSheet()
    H = ParagraphStyle("H", parent=ss["Title"], fontName="Helvetica-Bold",
                       fontSize=15, textColor=INK, spaceAfter=2)
    S = ParagraphStyle("S", parent=ss["Normal"], fontSize=9,
                       textColor=GREY, spaceAfter=6)
    CE = ParagraphStyle("CE", parent=ss["Normal"], fontSize=7.6,
                        leading=9.8, textColor=INK)
    H2 = ParagraphStyle("H2", parent=ss["Normal"], fontName="Helvetica-Bold",
                        fontSize=11, textColor=colors.HexColor("#c0392b"),
                        spaceBefore=12, spaceAfter=4)

    def hc(t):
        return Paragraph(f"<b>{t}</b>", ParagraphStyle(
            "hh", parent=CE, textColor=colors.white))

    page = landscape(A4)
    doc = SimpleDocTemplate(out_path, pagesize=page,
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=10*mm,
                            title=f"Customer Performance {report['month']}")
    avail = page[0] - 20*mm
    sofar = (f" (so far — {report['days_elapsed']}/{report['month_days']} days)"
             if report["partial"] else "")
    story = [Paragraph(f"Customer Performance — {report['month']}{sofar} · "
                       f"{report['scope']}", H),
             Paragraph(f"{report['customer_count']:,} customers · "
                       f"{report['not_ordered_count']} not ordered this "
                       f"month · generated {report['generated_at']}"
                       + (" · Diff/Change compared with the same number of "
                          "days last month" if report["partial"] else ""), S)]

    heads = [hc("Customer")] + [hc(k) for k in report["hist_keys"]]
    heads += [hc(f"{report['month']}{' so far' if report['partial'] else ''}")]
    if report["partial"]:
        heads += [hc("LM same days")]
    heads += [hc("Diff"), hc("Change"), hc("")]
    data = [heads]

    for r in report["customers"][:max_rows]:
        row = [Paragraph(r["customer"], CE)]
        row += [Paragraph(f"${v:,.0f}" if v else "-", CE)
                for v in r["history"]]
        row.append(Paragraph(f"<b>${r['this_month']:,.0f}</b>"
                             if r["this_month"] else "-", CE))
        if report["partial"]:
            row.append(Paragraph(
                f"${r['lm_same_days']:,.0f}" if r["lm_same_days"] else "-",
                CE))
        col = "#1e8449" if r["diff"] >= 0 else "#c0392b"
        row.append(Paragraph(
            f"<font color='{col}'>{'+' if r['diff'] >= 0 else ''}"
            f"${r['diff']:,.0f}</font>", CE))
        cp = r["change_pct"]
        row.append(Paragraph(
            f"<font color='{col}'><b>{'+' if (cp or 0) >= 0 else ''}"
            f"{cp}%</b></font>" if cp is not None else "-", CE))
        row.append(Paragraph(
            "<font color='#c0392b'><b>NOT ORDERED</b></font>"
            if r["not_ordered"] else "", CE))
        data.append(row)

    n_num = len(heads) - 2
    cw = [avail*0.26] + [avail*0.74/n_num]*(n_num) + [avail*0.0]
    cw = [avail*0.24] + [(avail*0.76)/(len(heads)-1)]*(len(heads)-1)
    t = Table(data, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("GRID", (0, 0), (-1, -1), 0.4, LINE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.6),
        *[("BACKGROUND", (0, i), (-1, i), SOFT)
          for i in range(2, len(data), 2)]
    ]))
    story.append(t)

    if report["not_ordered_top"]:
        story.append(Paragraph(
            f"Not ordered this month — top {len(report['not_ordered_top'])} "
            f"by last month's sales (call list)", H2))
        for i, r in enumerate(report["not_ordered_top"], 1):
            story.append(Paragraph(
                f"{i}. <b>{r['customer']}</b> — last month "
                f"${r['last_month']:,.2f}", CE))
    doc.build(story)
    return out_path


def customer_performance_xlsx(report: dict, out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1A2340")
    red_font = Font(bold=True, color="C0392B")
    thin = Border(*[Side(style="thin", color="D5DDE5")]*4)
    money = '"$"#,##0.00'
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Performance"
    sofar = (f" (so far - {report['days_elapsed']}/{report['month_days']} days)"
             if report["partial"] else "")
    ws["A1"] = f"Customer Performance — {report['month']}{sofar} · {report['scope']}"
    ws["A1"].font = Font(bold=True, size=13)
    if report["partial"]:
        ws["A2"] = "Diff/Change compared with the same number of days last month"
        ws["A2"].font = Font(size=9, color="5A6379")
    heads = (["Customer"] + report["hist_keys"]
             + [f"{report['month']} so far" if report["partial"]
                else report["month"]])
    if report["partial"]:
        heads += ["LM same days"]
    heads += ["Diff", "Change %", "Status"]
    for ci, h in enumerate(heads, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = thin
    r = 4
    for d in report["customers"]:
        r += 1
        vals = [d["customer"]] + d["history"] + [d["this_month"]]
        fmts = [None] + [money]*(len(d["history"]) + 1)
        if report["partial"]:
            vals.append(d["lm_same_days"]); fmts.append(money)
        vals += [d["diff"],
                 (d["change_pct"] / 100) if d["change_pct"] is not None
                 else None,
                 "NOT ORDERED" if d["not_ordered"] else ""]
        fmts += [money, '+0.0%;-0.0%', None]
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.border = thin
            if f and v is not None:
                cell.number_format = f
            if ci == len(vals) and v == "NOT ORDERED":
                cell.font = red_font
    ws.column_dimensions["A"].width = 46
    for ci in range(2, len(heads) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = "B5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(heads))}{r}"
    wb.save(out_path)
    return out_path


# ────────────────────────────────────────────────────────────────────
# Master data: customers and products (from Focus "Master info" exports)
# ────────────────────────────────────────────────────────────────────

def _find_header(raw, must_have=("name", "code")):
    for i in range(min(15, len(raw))):
        vals = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
        if all(any(m == v for v in vals) for m in must_have):
            return i
    raise ValueError("Header row not found — is this a Master info export?")


def _mcol(df, *cands):
    low = {str(col).strip().lower(): col for col in df.columns}
    for cand in cands:
        if cand in low:
            return low[cand]
    for cand in cands:
        for k, orig in low.items():
            if cand in k:
                return orig
    return None


def _mval(row, col):
    if col is None:
        return ""
    v = row.get(col)
    s = str(v if v is not None else "").replace("\xa0", " ").strip()
    return "" if s.lower() == "nan" else s


def parse_customers_master(data: bytes) -> tuple[list[dict], dict]:
    raw = pd.read_excel(io.BytesIO(data), sheet_name=0, header=None)
    h = _find_header(raw)
    df = raw.iloc[h + 1:].copy()
    df.columns = [str(x).strip() for x in raw.iloc[h].tolist()]

    col = {
        "name": _mcol(df, "name"),
        "code": _mcol(df, "code"),
        "acc_type": _mcol(df, "account account type", "account type"),
        "segment": _mcol(df, "customer segment name", "segment"),
        "area": _mcol(df, "deliver to area name", "area name"),
        "contact": _mcol(df, "contact person name"),
        "chain": _mcol(df, "chain store of name", "chain store"),
        "address": _mcol(df, "account billing address", "billing address"),
        "postal": _mcol(df, "account postal code", "postal code"),
        "mobile": _mcol(df, "account mobile", "mobile"),
        "whatsapp": _mcol(df, "account whatsapp number", "whatsapp"),
        "roc": _mcol(df, "account roc no", "roc no"),
        "modified": _mcol(df, "modified date"),
    }
    if not col["name"]:
        raise ValueError("'Name' column missing")

    rows, skipped = [], 0
    for _, r in df.iterrows():
        name = _mval(r, col["name"])
        if not name:
            continue
        acc_type = _mval(r, col["acc_type"])
        if acc_type.lower() != "customer":
            skipped += 1
            continue
        rows.append({
            "name": name,
            "code": _mval(r, col["code"]),
            "segment": _mval(r, col["segment"]),
            "area": _mval(r, col["area"]),
            "contact": _mval(r, col["contact"]),
            "chain_store": _mval(r, col["chain"]),
            "address": _mval(r, col["address"]),
            "postal_code": _mval(r, col["postal"]),
            "mobile": _mval(r, col["mobile"]),
            "whatsapp": _mval(r, col["whatsapp"]),
            "roc_no": _mval(r, col["roc"]),
            "modified_date": _mval(r, col["modified"]),
        })
    if not rows:
        raise ValueError("No Customer-type accounts found in the file")
    stats = {"customers": len(rows), "non_customer_accounts": skipped,
             "segments": len({r["segment"] for r in rows if r["segment"]}),
             "areas": len({r["area"] for r in rows if r["area"]}),
             "with_mobile": sum(1 for r in rows if r["mobile"])}
    return rows, stats


def upsert_customers_master(conn: sqlite3.Connection,
                            rows: list[dict]) -> dict:
    conn.executemany("""
        INSERT INTO customers_master
            (name, code, segment, area, contact, chain_store, address,
             postal_code, mobile, whatsapp, roc_no, modified_date)
        VALUES (:name,:code,:segment,:area,:contact,:chain_store,:address,
                :postal_code,:mobile,:whatsapp,:roc_no,:modified_date)
        ON CONFLICT(name) DO UPDATE SET
            code=excluded.code, segment=excluded.segment,
            area=excluded.area, contact=excluded.contact,
            chain_store=excluded.chain_store, address=excluded.address,
            postal_code=excluded.postal_code, mobile=excluded.mobile,
            whatsapp=excluded.whatsapp, roc_no=excluded.roc_no,
            modified_date=excluded.modified_date
    """, rows)
    conn.commit()
    return {"upserted": len(rows)}


def parse_products_master(data: bytes) -> tuple[list[dict], dict]:
    raw = pd.read_excel(io.BytesIO(data), sheet_name=0, header=None)
    h = _find_header(raw)
    df = raw.iloc[h + 1:].copy()
    df.columns = [str(x).strip() for x in raw.iloc[h].tolist()]

    col = {
        "name": _mcol(df, "name"),
        "code": _mcol(df, "code"),
        "supplier": _mcol(df, "supplier master name", "supplier"),
        "brand": _mcol(df, "brand master name", "brand"),
        "category": _mcol(df, "category master name"),
        "sub_category": _mcol(df, "sub category master name"),
        "sub_category2": _mcol(df, "sub category 2 master name"),
        "base_unit": _mcol(df, "default base unit name", "base unit"),
        "group_name": _mcol(df, "group master name"),
        "item_type": _mcol(df, "item item type", "item type"),
        "qty_ctn": _mcol(df, "item qty in ctn", "qty in ctn"),
        "selling_rate": _mcol(df, "item selling rate", "selling rate"),
        "min_sale": _mcol(df, "item minsaleprice", "minsaleprice"),
        "buying": _mcol(df, "item buyingprice", "buyingprice"),
        "modified": _mcol(df, "modified date"),
    }
    if not col["name"]:
        raise ValueError("'Name' column missing")

    rows = []
    for _, r in df.iterrows():
        name = _mval(r, col["name"])
        if not name:
            continue
        rows.append({
            "name": name,
            "code": _mval(r, col["code"]),
            "supplier": _mval(r, col["supplier"]),
            "brand": _mval(r, col["brand"]),
            "category": _mval(r, col["category"]),
            "sub_category": _mval(r, col["sub_category"]),
            "sub_category2": _mval(r, col["sub_category2"]),
            "base_unit": _mval(r, col["base_unit"]),
            "group_name": _mval(r, col["group_name"]),
            "item_type": _mval(r, col["item_type"]),
            "qty_per_ctn": _num(r.get(col["qty_ctn"])) if col["qty_ctn"] else 0,
            "selling_rate": _num(r.get(col["selling_rate"]))
                            if col["selling_rate"] else 0,
            "min_sale_price": _num(r.get(col["min_sale"]))
                              if col["min_sale"] else 0,
            "buying_price_ref": _num(r.get(col["buying"]))
                                if col["buying"] else 0,
            "modified_date": _mval(r, col["modified"]),
        })
    if not rows:
        raise ValueError("No items found in the file")
    stats = {"items": len(rows),
             "brands": len({r["brand"] for r in rows if r["brand"]}),
             "categories": len({r["category"] for r in rows if r["category"]}),
             "with_buying_price": sum(1 for r in rows
                                      if r["buying_price_ref"])}
    return rows, stats


def upsert_products_master(conn: sqlite3.Connection,
                           rows: list[dict]) -> dict:
    conn.executemany("""
        INSERT INTO products_master
            (name, code, supplier, brand, category, sub_category,
             sub_category2, base_unit, group_name, item_type, qty_per_ctn,
             selling_rate, min_sale_price, buying_price_ref, modified_date)
        VALUES (:name,:code,:supplier,:brand,:category,:sub_category,
                :sub_category2,:base_unit,:group_name,:item_type,
                :qty_per_ctn,:selling_rate,:min_sale_price,
                :buying_price_ref,:modified_date)
        ON CONFLICT(name) DO UPDATE SET
            code=excluded.code, supplier=excluded.supplier,
            brand=excluded.brand, category=excluded.category,
            sub_category=excluded.sub_category,
            sub_category2=excluded.sub_category2,
            base_unit=excluded.base_unit, group_name=excluded.group_name,
            item_type=excluded.item_type, qty_per_ctn=excluded.qty_per_ctn,
            selling_rate=excluded.selling_rate,
            min_sale_price=excluded.min_sale_price,
            buying_price_ref=excluded.buying_price_ref,
            modified_date=excluded.modified_date
    """, rows)
    conn.commit()
    return {"upserted": len(rows)}


# ────────────────────────────────────────────────────────────────────
# Price check: app pricebook (source of truth) vs Focus reference price
# ────────────────────────────────────────────────────────────────────

def build_price_check(conn: sqlite3.Connection,
                      threshold_pct: float = 2.0) -> dict:
    """Items where the app pricebook and the Focus item-master reference
    price disagree by more than threshold_pct. The pricebook remains the
    source of truth — this only flags candidates for review."""
    rows = conn.execute("""
        SELECT pb.item_name, pb.buying_price, pm.buying_price_ref,
               pm.brand, pm.supplier
        FROM pricebook pb
        JOIN products_master pm
          ON LOWER(TRIM(pm.name)) = LOWER(TRIM(pb.item_name))
        WHERE pb.buying_price > 0 AND pm.buying_price_ref > 0
    """).fetchall()
    out = []
    for name, app_p, ref_p, brand, supplier in rows:
        diff_pct = round((ref_p - app_p) / app_p * 100, 1)
        if abs(diff_pct) >= threshold_pct:
            out.append({"item": name, "pricebook": round(app_p, 4),
                        "focus_ref": round(ref_p, 4),
                        "diff_pct": diff_pct,
                        "brand": brand or "", "supplier": supplier or ""})
    out.sort(key=lambda x: -abs(x["diff_pct"]))
    not_in_master = conn.execute("""
        SELECT COUNT(*) FROM pricebook pb
        WHERE NOT EXISTS (SELECT 1 FROM products_master pm
                          WHERE LOWER(TRIM(pm.name)) =
                                LOWER(TRIM(pb.item_name)))
    """).fetchone()[0]
    return {"compared": len(rows), "flagged": out,
            "threshold_pct": threshold_pct,
            "pricebook_items_not_in_master": not_in_master,
            "generated_at": datetime.now().strftime("%d %b %Y, %I:%M %p")}
