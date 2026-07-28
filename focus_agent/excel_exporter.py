"""
excel_exporter.py
Creates a formatted Excel backup of the low-price alerts.
One sheet per day, all SOs with full item breakdown.
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


def export_alerts(alerts: list, output_dir: str = "downloads") -> str:
    os.makedirs(output_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    filepath = os.path.join(output_dir, f"LowPrice_Alert_{date_str}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Alerts {date_str}"

    # ── Colour palette ────────────────────────────────────────────────────
    RED    = "C0392B"
    ORANGE = "E67E22"
    GREEN  = "27AE60"
    HEADER = "1A252F"
    SO_ROW = "2C3E50"
    ALT    = "EBF5FB"
    WHITE  = "FFFFFF"

    def cell_style(cell, bold=False, fg=None, bg=None,
                   align="left", border=False, num_fmt=None):
        cell.font      = Font(name="Arial", bold=bold,
                              color=fg or "000000", size=10)
        if bg:
            cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center",
                                   wrap_text=False)
        if border:
            thin = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=thin, right=thin,
                                 top=thin, bottom=thin)
        if num_fmt:
            cell.number_format = num_fmt

    # ── Title row ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value = f"🚨 Low Price Alert — {datetime.now().strftime('%d %b %Y %I:%M %p')}"
    cell_style(title, bold=True, fg=WHITE, bg=HEADER, align="center")
    ws.row_dimensions[1].height = 28

    # ── Column headers ─────────────────────────────────────────────────────
    headers = [
        "SO Number", "Customer", "Salesman", "Item",
        "Unit", "Pcs/CTN", "Qty Ordered",
        "Cost/pc ($)", "Landing/pc ($)", "Prev Price/pc ($)",
        "Rate/pc ($)", "Margin %"
    ]
    col_widths = [14, 32, 16, 40, 8, 8, 12, 12, 13, 14, 10, 10]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        cell_style(c, bold=True, fg=WHITE, bg="2E4053", align="center", border=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    # ── Data rows ──────────────────────────────────────────────────────────
    row = 3
    for alert in alerts:
        voucher  = alert["voucher"]
        party    = alert["party"]
        salesman = alert.get("salesman", "")

        for i, item in enumerate(alert["items"]):
            margin = item["margin"]

            # Margin colour
            if margin < 0:
                margin_bg = "FADBD8"   # red tint
            elif margin < 5:
                margin_bg = "FDEBD0"   # orange tint
            else:
                margin_bg = "D5F5E3"   # green tint

            bg = ALT if i % 2 == 0 else WHITE

            values = [
                voucher if i == 0 else "",
                party   if i == 0 else "",
                salesman if i == 0 else "",
                item["item"],
                item["unit"].upper().lstrip("."),
                item["pieces_per_unit"],
                item["quantity"],
                item["pricebook"],
                item["landing"],
                item["prev_price"],
                item["rate_per_piece"],
                margin / 100,
            ]

            num_fmts = [
                None, None, None, None, None,
                "#,##0", "#,##0",
                '$#,##0.00', '$#,##0.00', '$#,##0.00',
                '$#,##0.00',
                '0.0%'
            ]

            for col, (val, fmt) in enumerate(zip(values, num_fmts), 1):
                c = ws.cell(row=row, column=col, value=val)
                use_bg = margin_bg if col == 12 else bg
                cell_style(c, bg=use_bg, border=True, num_fmt=fmt,
                           align="right" if col >= 6 else "left")

            ws.row_dimensions[row].height = 16
            row += 1

        # Blank separator row between SOs
        row += 1

    # ── Summary at bottom ──────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="Total flagged SOs:").font = Font(bold=True, name="Arial")
    ws.cell(row=row, column=2, value=len(alerts)).font = Font(bold=True, name="Arial")

    total_items = sum(len(a["items"]) for a in alerts)
    ws.cell(row=row+1, column=1, value="Total flagged items:").font = Font(bold=True, name="Arial")
    ws.cell(row=row+1, column=2, value=total_items).font = Font(bold=True, name="Arial")

    wb.save(filepath)
    print(f"   📊 Excel saved: {filepath}")
    return filepath
